"""시나리오 분석 결과를 Excel 로 내보내기."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E5077", end_color="2E5077", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(vertical="center", wrap_text=True)


def export_scenario_excel(
    title: str,
    characters: list,
    locations: list,
    props: list,
    fx: list,
    shots: list,
    dialogues: list,
    storyboard_dir: Path | None,
    output_path: Path,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # ── 캐릭터 ────────────────────────────────────────
    _make_sheet(
        wb,
        "캐릭터",
        ["이름", "분류", "묘사", "비고", "등장 샷 수", "등장 샷 코드"],
        [
            [
                c.get("name", ""),
                c.get("category", ""),
                c.get("description", ""),
                c.get("notes", ""),
                c.get("appearance_count", 0),
                ", ".join(c.get("shot_codes", []) or []),
            ]
            for c in (characters or [])
        ],
        widths=[20, 14, 40, 24, 10, 40],
    )

    # ── 장소 ──────────────────────────────────────────
    _make_sheet(
        wb,
        "장소",
        ["이름", "분류", "시간대", "묘사", "등장 샷 수", "등장 샷 코드"],
        [
            [
                l.get("name", ""),
                l.get("category", ""),
                l.get("time_of_day", ""),
                l.get("description", ""),
                l.get("appearance_count", 0),
                ", ".join(l.get("shot_codes", []) or []),
            ]
            for l in (locations or [])
        ],
        widths=[20, 12, 14, 40, 10, 40],
    )

    # ── 소품/에셋 ─────────────────────────────────────
    _make_sheet(
        wb,
        "소품_에셋",
        ["이름", "분류", "묘사", "등장 샷 수", "등장 샷 코드"],
        [
            [
                p.get("name", ""),
                p.get("category", ""),
                p.get("description", ""),
                p.get("appearance_count", 0),
                ", ".join(p.get("shot_codes", []) or []),
            ]
            for p in (props or [])
        ],
        widths=[20, 14, 40, 10, 40],
    )

    # ── FX ────────────────────────────────────────────
    _make_sheet(
        wb,
        "FX",
        ["이름", "분류", "묘사", "등장 샷 수", "등장 샷 코드"],
        [
            [
                f.get("name", ""),
                f.get("category", ""),
                f.get("description", ""),
                f.get("appearance_count", 0),
                ", ".join(f.get("shot_codes", []) or []),
            ]
            for f in (fx or [])
        ],
        widths=[20, 14, 40, 10, 40],
    )

    # ── 샷 리스트 (스토리보드 이미지 포함) ────────────
    _make_shots_sheet(wb, shots or [], storyboard_dir)

    # ── 대사 ──────────────────────────────────────────
    _make_sheet(
        wb,
        "대사",
        ["씬", "캐릭터", "대사"],
        [[d.get("scene_number", ""), d.get("character", ""), d.get("line", "")] for d in (dialogues or [])],
        widths=[10, 16, 60],
    )

    # ── 첫 시트로 샷 리스트가 보이도록 순서 조정 ──────
    if "샷리스트" in wb.sheetnames:
        wb.move_sheet("샷리스트", offset=-(len(wb.sheetnames) - wb.sheetnames.index("샷리스트")))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _make_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list], widths: list[int]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append(r)
        ri = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=ri, column=c).alignment = BODY_ALIGN


SHOT_HEADERS = [
    "샷#",
    "코드",
    "씬",
    "샷사이즈",
    "카메라무빙",
    "카메라앵글",
    "렌즈",
    "시간대",
    "조명",
    "캐릭터",
    "장소",
    "소품",
    "FX(사용)",
    "액션/연기",
    "대사",
    "FX",
    "비고",
    "스토리보드",
]
SHOT_WIDTHS = [6, 16, 8, 10, 12, 14, 10, 10, 20, 22, 18, 20, 20, 30, 26, 18, 22, 28]
SHOT_THUMB_COL = 18
SHOT_THUMB_W = 200
SHOT_THUMB_H = 113  # 16:9
SHOT_ROW_HEIGHT = 90


def _make_shots_sheet(wb: Workbook, shots: list, storyboard_dir: Path | None) -> None:
    ws = wb.create_sheet("샷리스트")
    ws.append(SHOT_HEADERS)
    for col_idx in range(1, len(SHOT_HEADERS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
    for i, w in enumerate(SHOT_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for i, shot in enumerate(shots):
        chars = shot.get("characters") or []
        chars_str = ", ".join(chars) if isinstance(chars, list) else str(chars)
        props_used = shot.get("props_used") or []
        props_str = ", ".join(props_used) if isinstance(props_used, list) else str(props_used)
        fx_used = shot.get("fx_used") or []
        fx_used_str = ", ".join(fx_used) if isinstance(fx_used, list) else str(fx_used)
        seq = shot.get("sequence_number")
        sn = shot.get("shot_number")
        code = f"S{seq:04d}_C{sn:04d}" if isinstance(seq, int) and isinstance(sn, int) else ""
        row = [
            i + 1,
            code,
            shot.get("scene_number", ""),
            shot.get("shot_size", ""),
            shot.get("camera_movement", ""),
            shot.get("camera_angle", ""),
            shot.get("lens_mm", ""),
            shot.get("time_of_day", ""),
            shot.get("lighting", ""),
            chars_str,
            shot.get("location", ""),
            props_str,
            fx_used_str,
            shot.get("action", ""),
            shot.get("dialogue", "") or "",
            shot.get("fx", "") or "",
            shot.get("notes", "") or "",
            "",  # 이미지 셀
        ]
        ws.append(row)
        ri = ws.max_row
        ws.row_dimensions[ri].height = SHOT_ROW_HEIGHT
        for c in range(1, len(SHOT_HEADERS) + 1):
            ws.cell(row=ri, column=c).alignment = BODY_ALIGN

        if storyboard_dir is not None:
            img_path = storyboard_dir / f"shot_{i:04d}.png"
            if img_path.exists():
                try:
                    img = XLImage(str(img_path))
                    img.width = SHOT_THUMB_W
                    img.height = SHOT_THUMB_H
                    anchor = f"{get_column_letter(SHOT_THUMB_COL)}{ri}"
                    ws.add_image(img, anchor)
                except Exception:
                    pass
