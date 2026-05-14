"""Excel/CSV 출력."""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Shot, ShotAnalysis, format_shot_code


HEADERS = [
    "컷#",
    "코드",
    "시작 TC",
    "끝 TC",
    "길이(초)",
    "길이(프레임)",
    "썸네일",
    "샷사이즈",
    "카메라 무빙",
    "캐릭터",
    "배경",
    "액션/연기",
    "대사",
    "FX",
    "비고",
]

COLUMN_WIDTHS = [6, 16, 14, 14, 9, 11, 22, 10, 14, 26, 28, 32, 26, 18, 22]
THUMB_COL_INDEX = 7  # G열 (1-base, "코드" 컬럼이 앞에 추가됨)
THUMB_WIDTH_PX = 140
THUMB_HEIGHT_PX = 80
ROW_HEIGHT_PT = 64


def export_excel(shots: list[Shot], output_path: Path) -> Path:
    """샷 리스트를 Excel(.xlsx)로 저장. 썸네일은 F열에 임베딩."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shot List"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E5077", end_color="2E5077", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, w in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    body_align = Alignment(vertical="center", wrap_text=True)

    for shot in shots:
        a = shot.analysis or ShotAnalysis()
        row = [
            shot.index,
            format_shot_code(shot.sequence_number, shot.shot_number),
            shot.start_tc,
            shot.end_tc,
            f"{shot.duration_seconds:.2f}",
            shot.duration_frames,
            "",  # 썸네일 셀 (이미지로 채워짐)
            a.shot_size or "",
            a.camera_movement or "",
            ", ".join(a.characters or []),
            a.background or "",
            a.action or "",
            shot.dialogue or "",
            a.fx or "",
            a.notes or "",
        ]
        ws.append(row)
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).alignment = body_align

        if shot.thumbnail_path and shot.thumbnail_path.exists():
            img = XLImage(str(shot.thumbnail_path))
            img.width = THUMB_WIDTH_PX
            img.height = THUMB_HEIGHT_PX
            anchor = f"{get_column_letter(THUMB_COL_INDEX)}{row_idx}"
            ws.add_image(img, anchor)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_csv(shots: list[Shot], output_path: Path) -> Path:
    """CSV 출력 (썸네일 경로만 텍스트로 기록)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS + ["썸네일 경로"])
        for shot in shots:
            a = shot.analysis or ShotAnalysis()
            writer.writerow(
                [
                    shot.index,
                    format_shot_code(shot.sequence_number, shot.shot_number),
                    shot.start_tc,
                    shot.end_tc,
                    f"{shot.duration_seconds:.2f}",
                    shot.duration_frames,
                    "",
                    a.shot_size or "",
                    a.camera_movement or "",
                    ", ".join(a.characters or []),
                    a.background or "",
                    a.action or "",
                    shot.dialogue or "",
                    a.fx or "",
                    a.notes or "",
                    str(shot.thumbnail_path) if shot.thumbnail_path else "",
                ]
            )
    return output_path
